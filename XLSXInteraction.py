import openpyxl
import os
import datetime
import time
#define file path
path = os.getcwd()
#Make a directory of files
files = []
for r, d, f in os.walk(path):
    for file in f:
        if('.xlsx' in file or '.xls' in file):
            files.append(file)
print(files)
#read in the template spreadsheet
wb_new = openpyxl.load_workbook(files[0])
sheets_new = wb_new.get_sheet_names()
sheet_new = wb_new.get_sheet_by_name(sheets_new[0])

#Need to count which row of the new sheet we are dealing with
count = 2
batch = ''
t_downs = [1]
#sheet, check batch
for ii in range (2,len(files)):
    wb = openpyxl.load_workbook(files[ii])
    sheets = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheets[5])
    if(batch != sheet.cell(row=3,column=3).value):
        batch = sheet.cell(row=3,column=3).value
        t_downs = [1]
    jj = 25
    flag = False
    #each of these is for a certain row
    while (not flag):
        flag_dupe_t = False
        name = sheet.cell(row=jj,column=1).value
        t_down = sheet.cell(row=jj,column=2).value
        t_up = sheet.cell(row=jj,column=3).value
        comment = sheet.cell(row=jj,column=6).value
        if(not sheet.cell(row=jj+1,column=1).value):
            flag =True
        jj += 1
        if(name == 'Comment' or name == 'Prior to Fill Samples'):
            event = comment
        else:
            event = name
        for qq in range (0,len(t_downs)):
            if(t_down == t_downs[qq]) and (t_up):
                flag_dupe_t = True
        if(t_up and isinstance(t_up,datetime.time)):
            t_downs.append(t_down)
        if (event):
            event = event.replace('.','')
            event_splt = event.split()
        shift = 0
        lack_flag = False
        needle = 0
        for tt in range (0, len(event_splt)):
            if event_splt[tt] == 'Shift' or event_splt[tt] == 'shift':
                shift += 1
            if event_splt[tt] == 'Change' or event_splt[tt] == 'change':
                shift += 1
            if event_splt[tt] == 'plating' or event_splt[tt] == 'Plating':
                event = 'Personnel Plating'
            if event_splt[tt] == 'waiting' or event_splt[tt] == 'Waiting' or event_splt[tt] == 'components' or event_splt[tt] =='Components':
                event = 'Waiting on components'
            if event_splt[tt] == 'needle' or event_splt[tt] == 'Needle':
                needle+=1
            if event_splt[tt] == 'stick' or event_splt[tt] == 'Stick':
                needle+=1
        if(needle == 2):
            event = 'Needle Stick'
        if (shift == 2):
            event = 'Shift Change'
        if (comment):
            comment = comment.replace('.','')
        if(t_up) and (t_down) and isinstance(t_up,datetime.time):
            hour_up = t_up.hour
            min_up = t_up.minute
            hour_down = t_down.hour
            min_down = t_down.minute
        if(hour_down!=None):
            if (hour_down > hour_up):
                hour_down = 23 - hour_down
                min_down = 60 - min_down
                hours = hour_down + hour_up
                mins = min_down+min_up
                if(mins>60):
                    hours +=1
                    mins -=60
                downtime = datetime.time(hours,mins)
                downtime2 = hours + mins/60
            else:
                hours = hour_up-hour_down
                mins = min_up-min_down
                if(mins>60):
                    hours+=1
                    mins-=60
                if(mins<0):
                    hours-= 1
                    mins += 60
                downtime = datetime.time(hours,mins)
                downtime2 = hours + mins/60
        if(not flag_dupe_t) and (t_up) and isinstance(t_up,datetime.time):
            sheet_new.cell(row=count,column=1).value = batch
            sheet_new.cell(row=count,column=2).value = downtime
            sheet_new.cell(row=count,column=3).value = downtime2
            sheet_new.cell(row=count,column=4).value = event
            sheet_new.cell(row=count,column=5).value = t_up
            sheet_new.cell(row=count,column=6).value = t_down
            sheet_new.cell(row=count,column=7).value = comment
            sheet_new.cell(row=count,column=8).value = ii
            sheet_new.cell(row=count,column=9).value = jj
            count+=1
wb_new.save('00Failure Modes.xlsx')

#Need to check