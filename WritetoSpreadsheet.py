# -*- coding: utf-8 -*-
"""
Created on Sun Sep 20 20:09:59 2020

@author: josh
This file takes in a JSON response about lead based product recalls, then creates
a spreadsheet that records each recall,separate by the response product labels

"""

from openpyxl import Workbook
import time
def WritetoSpreadsheet(response):
    #Create the new spreadsheet
    wb_new = Workbook()
    sheets_new = wb_new.get_sheet_names()
    sheet_new = wb_new.get_sheet_by_name(sheets_new[0])
    sheet_new.cell(row=1,column=1).value = 'Recall Date'
    sheet_new.cell(row=1,column=2).value = 'Recall ID'
    sheet_new.cell(row=1,column=3).value = 'Title'
    sheet_new.cell(row=1,column=4).value = 'Description'
    sheet_new.cell(row=1,column=5).value = 'Consumer Contact'
    sheet_new.cell(row=1,column=6).value = 'URL for Recall'
    sheet_new.cell(row=1,column=7).value = 'Product'
    sheet_new.cell(row=1,column=8).value = 'Injuries'
    sheet_new.cell(row=1,column=9).value = 'Retailers'
    sheet_new.cell(row=1,column=10).value = 'Distributors'
    sheet_new.cell(row=1,column=11).value = 'Importers'
    sheet_new.cell(row=1,column=12).value = 'Manufacturer Country'
    sheet_new.cell(row=1,column=13).value = 'Hazards'
    sheet_new.cell(row=1,column=14).value = 'Remedies'
    sheet_new.cell(row=1,column=15).value = 'Remedy Options'
    
    #Cycle through the response and write the spreadsheet for the responses
    for ii in range (0,len(response.json())):
        sheet_new.cell(row=ii+2,column=1).value = response.json()[ii]['RecallDate']
        sheet_new.cell(row=ii+2,column=2).value = response.json()[ii]['RecallID']
        sheet_new.cell(row=ii+2,column=3).value = response.json()[ii]['Title']
        sheet_new.cell(row=ii+2,column=4).value = response.json()[ii]['Description']
        sheet_new.cell(row=ii+2,column=5).value = response.json()[ii]['ConsumerContact']
        sheet_new.cell(row=ii+2,column=6).value = response.json()[ii]['URL']
        sheet_new.cell(row=ii+2,column=7).value = response.json()[ii]['Products'][0]['Name']
        try:
            sheet_new.cell(row=ii+2,column=8).value = response.json()[ii]['Injuries'][0]['Name']
        except:
            sheet_new.cell(row=ii+2,column=8).value = 'No Information'
        try:
            sheet_new.cell(row=ii+2,column=9).value = response.json()[ii]['Retailers'][0]['Name']
        except:
            sheet_new.cell(row=ii+2,column=9).value = 'No Information'
        try:
            sheet_new.cell(row=ii+2,column=10).value = response.json()[ii]['Distributors'][0]['Name']
        except:
            sheet_new.cell(row=ii+2,column=10).value = 'No information'
        try:
            sheet_new.cell(row=ii+2,column=11).value = response.json()[ii]['Importers'][0]['Name']
        except:
            sheet_new.cell(row=ii+2,column=11).value = 'No Information'
        try:
            sheet_new.cell(row=ii+2,column=12).value = response.json()[ii]['ManufacturerCountries'][0]['Country']
        except:
            sheet_new.cell(row=ii+2,column=13).value = 'No Information'
        try:
            sheet_new.cell(row=ii+2,column=13).value = response.json()[ii]['Hazards'][0]['Name']
        except:
            sheet_new.cell(row=ii+2,column=13).value = 'No Information'
        try:
            sheet_new.cell(row=ii+2,column=14).value = response.json()[ii]['Remedies'][0]['Name']
        except:
            sheet_new.cell(row=ii+2,column=14).value = 'No Information'
        try:
            sheet_new.cell(row=ii+2,column=15).value = response.json()[ii]['RemedyOptions'][0]['Option']
        except:
            sheet_new.cell(row=ii+2,column=15).value = 'No Information'
    #Save the spreadsheet of responses
    wb_new.save('Lead Related Product Recalls.xlsx')